"""
ProChain Daily Report System (Gmail Edition)
==============================================
อ่านอีเมล Gmail → ดาวน์โหลด Excel → บันทึก Google Drive → เขียน Google Sheet

Environment variables ที่ต้องตั้งใน GitHub Secrets:
  GMAIL_CLIENT_ID
  GMAIL_CLIENT_SECRET
  GMAIL_REFRESH_TOKEN
  GOOGLE_SERVICE_ACCOUNT_JSON   — สำหรับ Sheets + Drive (service account เดิม)
  SPREADSHEET_ID
  DRIVE_FOLDER_ID                — โฟลเดอร์หลักใน Google Drive สำหรับเก็บไฟล์ต้นฉบับ
  SYSTEM_ACTIVE                  — TRUE หรือ FALSE (dead key)
"""

import os
import io
import json
import base64
import logging
from datetime import datetime, timezone, timedelta

from google.oauth2.credentials import Credentials as UserCredentials
from google.oauth2.service_account import Credentials as ServiceCredentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
from openpyxl import load_workbook
import gspread

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

# ─── Config ────────────────────────────────────────────────────────
GMAIL_CLIENT_ID     = os.environ["GMAIL_CLIENT_ID"]
GMAIL_CLIENT_SECRET = os.environ["GMAIL_CLIENT_SECRET"]
GMAIL_REFRESH_TOKEN = os.environ["GMAIL_REFRESH_TOKEN"]
SPREADSHEET_ID       = os.environ["SPREADSHEET_ID"]
DRIVE_FOLDER_ID      = os.environ["DRIVE_FOLDER_ID"]
SYSTEM_ACTIVE        = os.environ.get("SYSTEM_ACTIVE", "TRUE").upper()

SUBJECT_KEYWORD = "daily report"

GMAIL_SCOPES = [
    "https://www.googleapis.com/auth/gmail.readonly",
    "https://www.googleapis.com/auth/drive",
]
SHEETS_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

# ─── Dead Key ──────────────────────────────────────────────────────
def check_dead_key():
    if SYSTEM_ACTIVE != "TRUE":
        log.warning("🔴 SYSTEM_ACTIVE = FALSE — ระบบถูกปิดการทำงาน ไม่บันทึกข้อมูล")
        return False
    return True

# ─── Gmail + Drive Auth (OAuth User) ───────────────────────────────
def get_user_creds():
    """ใช้ OAuth token เดียวกันสำหรับทั้ง Gmail และ Drive"""
    return UserCredentials(
        token=None,
        refresh_token=GMAIL_REFRESH_TOKEN,
        client_id=GMAIL_CLIENT_ID,
        client_secret=GMAIL_CLIENT_SECRET,
        token_uri="https://oauth2.googleapis.com/token",
        scopes=GMAIL_SCOPES,
    )

def get_gmail_service():
    return build("gmail", "v1", credentials=get_user_creds())

def get_drive_service():
    """ใช้ OAuth ของเจ้าของ Drive แทน Service Account เพื่อไม่ติด storage quota"""
    return build("drive", "v3", credentials=get_user_creds())

# ─── Google Sheets Auth (Service Account) ──────────────────────────
def get_service_account_creds():
    sa_json = os.environ["GOOGLE_SERVICE_ACCOUNT_JSON"]
    creds_dict = json.loads(sa_json)
    return ServiceCredentials.from_service_account_info(
        creds_dict, scopes=SHEETS_SCOPES
    )

def get_gsheet_client():
    creds = get_service_account_creds()
    return gspread.authorize(creds)

# ─── Gmail: ดึงอีเมลวันนี้ ─────────────────────────────────────────
def fetch_today_emails(service) -> list[dict]:
    tz_thai = timezone(timedelta(hours=7))
    today_str = datetime.now(tz_thai).strftime("%Y/%m/%d")

    query = f'subject:"{SUBJECT_KEYWORD}" has:attachment after:{today_str}'
    results = service.users().messages().list(userId="me", q=query).execute()
    messages = results.get("messages", [])
    log.info(f"📬 พบ {len(messages)} อีเมลวันนี้")
    return messages

# ─── Gmail: ดึง attachments ────────────────────────────────────────
def fetch_attachments(service, message_id: str) -> list[dict]:
    msg = service.users().messages().get(userId="me", id=message_id).execute()
    subject = ""
    for header in msg["payload"].get("headers", []):
        if header["name"].lower() == "subject":
            subject = header["value"]

    attachments = []
    parts = msg["payload"].get("parts", [])
    for part in parts:
        filename = part.get("filename", "")
        if filename.lower().endswith((".xlsx", ".xls")):
            att_id = part["body"].get("attachmentId")
            if att_id:
                att = service.users().messages().attachments().get(
                    userId="me", messageId=message_id, id=att_id
                ).execute()
                content_b64 = att["data"]
                # Gmail ใช้ URL-safe base64 ต้องแปลงก่อนใช้
                content_bytes = base64.urlsafe_b64decode(content_b64)
                content_std_b64 = base64.b64encode(content_bytes).decode()
                attachments.append({
                    "name": filename,
                    "content_b64": content_std_b64,
                })
    return attachments, subject

# ─── Google Drive: บันทึกไฟล์ ──────────────────────────────────────
def get_or_create_folder(drive_service, name: str, parent_id: str) -> str:
    query = (
        f"name='{name}' and '{parent_id}' in parents "
        f"and mimeType='application/vnd.google-apps.folder' and trashed=false"
    )
    results = drive_service.files().list(q=query, fields="files(id, name)").execute()
    files = results.get("files", [])
    if files:
        return files[0]["id"]

    metadata = {
        "name": name,
        "mimeType": "application/vnd.google-apps.folder",
        "parents": [parent_id],
    }
    folder = drive_service.files().create(body=metadata, fields="id").execute()
    return folder["id"]

def save_to_drive(drive_service, filename: str, content_b64: str, project_name: str, report_date: str):
    content = base64.b64decode(content_b64)

    try:
        date_obj = datetime.strptime(report_date.strip(), "%d/%m/%Y")
        month_folder_name = date_obj.strftime("%Y-%m")
    except Exception:
        month_folder_name = datetime.now().strftime("%Y-%m")

    project_folder_id = get_or_create_folder(drive_service, project_name or "Unknown Project", DRIVE_FOLDER_ID)
    month_folder_id = get_or_create_folder(drive_service, month_folder_name, project_folder_id)

    media = MediaIoBaseUpload(
        io.BytesIO(content),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    metadata = {"name": filename, "parents": [month_folder_id]}
    drive_service.files().create(body=metadata, media_body=media, fields="id").execute()
    log.info(f"💾 บันทึกไฟล์ Drive: {project_name}/{month_folder_name}/{filename}")

# ─── Excel: อ่านข้อมูล ────────────────────────────────────────────
def parse_excel(content_b64: str, filename: str) -> dict:
    content = base64.b64decode(content_b64)
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    def cell(r, c):
        v = ws.cell(row=r, column=c).value
        if v is None:
            return ""
        if isinstance(v, datetime):
            return v.strftime("%d/%m/%Y")
        return str(v).strip()

    project_name     = cell(7, 3)
    report_date      = cell(8, 3)
    week_no          = cell(8, 6)
    report_no        = cell(8, 8)   # H8
    site_manager     = cell(9, 3)
    contract_no      = cell(9, 6)
    work_hours       = cell(10, 3)
    weather_am       = cell(10, 7)
    weather_pm       = cell(10, 9)
    workforce        = cell(13, 8)
    accident         = cell(39, 2)
    near_miss        = cell(39, 3)
    first_aid        = cell(39, 4)
    safety_talk      = cell(39, 5)  # E39
    safety_status    = cell(39, 8)  # H39
    incident_desc    = cell(44, 1)  # A44
    imported_at      = datetime.now().strftime("%d/%m/%Y %H:%M")

    activities = []
    r = 16
    ACTIVITY_TABLE_END = 19  # tbl_Activity สิ้นสุดที่ row 19 (header row 15 + 4 data rows)
    while r <= ACTIVITY_TABLE_END:
        act_code = cell(r, 1)
        act_desc = cell(r, 3)
        plan_pct = cell(r, 6)
        actual_pct = cell(r, 7)
        remark   = cell(r, 8)

        if not act_code and not act_desc:
            r += 1
            continue

        activities.append({
            "วันที่": report_date, "ชื่อโครงการ": project_name, "Report No.": report_no,
            "ActivityCode": act_code, "Description": act_desc,
            "PlanPct": plan_pct, "ActualPct": actual_pct,
            "Variance": str(round(float(actual_pct) - float(plan_pct), 2)) if plan_pct and actual_pct else "",
            "Status": "Delayed" if plan_pct and actual_pct and float(actual_pct) < float(plan_pct) else "On Track",
            "Remark": remark, "ไฟล์ต้นฉบับ": filename, "นำเข้าเมื่อ": imported_at,
        })
        r += 1

    materials = []
    r = 22
    MATERIAL_TABLE_END = 29  # tbl_Material สิ้นสุดที่ row 29 (header row 21 + 8 data rows)
    while r <= MATERIAL_TABLE_END:
        mat_code = cell(r, 2)
        mat_desc = cell(r, 3)
        mat_unit = cell(r, 5)
        mat_qty  = cell(r, 6)
        mat_inout = cell(r, 7)
        mat_supplier = cell(r, 8)

        if not mat_desc and not mat_qty:
            r += 1
            continue

        materials.append({
            "วันที่": report_date, "ชื่อโครงการ": project_name, "Report No.": report_no,
            "MatCode": mat_code, "Description": mat_desc, "Unit": mat_unit,
            "Qty": mat_qty, "InOut": mat_inout, "Supplier": mat_supplier,
            "ไฟล์ต้นฉบับ": filename, "นำเข้าเมื่อ": imported_at,
        })
        r += 1

    daily_log = {
        "วันที่": report_date, "ชื่อโครงการ": project_name, "Report No.": report_no,
        "สัปดาห์ที่": week_no, "Site Manager": site_manager, "เลขที่สัญญา": contract_no,
        "ชั่วโมงทำงาน": work_hours, "อากาศเช้า": weather_am, "อากาศบ่าย": weather_pm,
        "คนงานรวม": workforce, "อุบัติเหตุ": accident, "Near Miss": near_miss,
        "First Aid": first_aid, "Safety Talk": safety_talk, "Safety Status": safety_status,
        "Incident Description": incident_desc,
        "นำเข้าเมื่อ": imported_at, "ไฟล์ต้นฉบับ": filename,
    }

    log.info(f"📊 อ่านข้อมูล: {len(activities)} กิจกรรม, {len(materials)} รายการวัสดุ")
    return {
        "project_name": project_name, "report_date": report_date,
        "daily_log": daily_log, "activities": activities, "materials": materials,
    }

# ─── Google Sheets ────────────────────────────────────────────────
def check_system_active_in_sheet(ss) -> bool:
    try:
        config = ss.worksheet("Config")
        val = config.acell("B1").value
        if str(val).upper() != "TRUE":
            log.warning("🔴 Config sheet: SYSTEM_ACTIVE = FALSE — หยุดการบันทึก")
            return False
        return True
    except Exception as e:
        log.warning(f"⚠️ อ่าน Config sheet ไม่ได้: {e} — ดำเนินการต่อ")
        return True

def append_to_sheet(ss, sheet_name: str, rows: list[list]):
    ws = ss.worksheet(sheet_name)
    if rows:
        ws.append_rows(rows, value_input_option="USER_ENTERED")
        log.info(f"✅ บันทึก {len(rows)} แถวลง '{sheet_name}'")

def update_stock_summary(ss, materials: list[dict]):
    if not materials:
        return
    ws = ss.worksheet("Stock Summary")
    existing = ws.get_all_values()
    existing_keys = set()
    for row in existing[1:]:
        if len(row) >= 2:
            existing_keys.add((row[0], row[1]))

    new_rows = []
    for m in materials:
        key = (m["ชื่อโครงการ"], m["MatCode"])
        if key not in existing_keys:
            row_num = len(existing) + len(new_rows) + 1
            new_rows.append([
                m["ชื่อโครงการ"], m["MatCode"], m["Description"], m["Unit"],
                f'=SUMIFS(\'Material Ledger\'!G:G,\'Material Ledger\'!B:B,A{row_num},\'Material Ledger\'!D:D,B{row_num},\'Material Ledger\'!H:H,"IN")',
                f'=SUMIFS(\'Material Ledger\'!G:G,\'Material Ledger\'!B:B,A{row_num},\'Material Ledger\'!D:D,B{row_num},\'Material Ledger\'!H:H,"OUT")',
                f'=E{row_num}-F{row_num}',
                "",
                f'=MAXIFS(\'Material Ledger\'!A:A,\'Material Ledger\'!B:B,A{row_num},\'Material Ledger\'!D:D,B{row_num})',
            ])
            existing_keys.add(key)

    if new_rows:
        ws.append_rows(new_rows, value_input_option="USER_ENTERED")
        log.info(f"✅ เพิ่ม {len(new_rows)} รายการใหม่ใน Stock Summary")

def is_duplicate(ss, report_no: str, project_name: str) -> bool:
    """เช็คว่า Report No. + ชื่อโครงการ นี้มีใน Daily Log แล้วหรือยัง"""
    try:
        ws = ss.worksheet("Daily Log")
        values = ws.get_all_values()
        for row in values[1:]:  # skip header
            if len(row) >= 3 and row[2] == report_no and row[1] == project_name:
                log.warning(f"⚠️ พบข้อมูลซ้ำ Report No.={report_no} โครงการ={project_name} — ข้ามการบันทึก")
                return True
        return False
    except Exception:
        return False

def write_to_gsheet(gc, data: dict):
    ss = gc.open_by_key(SPREADSHEET_ID)

    if not check_system_active_in_sheet(ss):
        return

    dl = data["daily_log"]

    # เช็คซ้ำก่อนบันทึก
    if is_duplicate(ss, dl["Report No."], dl["ชื่อโครงการ"]):
        return
    daily_row = [[
        dl["วันที่"], dl["ชื่อโครงการ"], dl["Report No."], dl["สัปดาห์ที่"],
        dl["Site Manager"], dl["เลขที่สัญญา"], dl["ชั่วโมงทำงาน"],
        dl["อากาศเช้า"], dl["อากาศบ่าย"], dl["คนงานรวม"],
        dl["อุบัติเหตุ"], dl["Near Miss"], dl["First Aid"],
        dl["Safety Talk"], dl["Safety Status"], dl["Incident Description"],
        dl["นำเข้าเมื่อ"], dl["ไฟล์ต้นฉบับ"],
    ]]
    append_to_sheet(ss, "Daily Log", daily_row)

    act_rows = [[
        a["วันที่"], a["ชื่อโครงการ"], a["Report No."], a["ActivityCode"], a["Description"],
        a["PlanPct"], a["ActualPct"], a["Variance"], a["Status"], a["Remark"],
        a["ไฟล์ต้นฉบับ"], a["นำเข้าเมื่อ"],
    ] for a in data["activities"]]
    append_to_sheet(ss, "Work Progress", act_rows)

    mat_rows = [[
        m["วันที่"], m["ชื่อโครงการ"], m["Report No."], m["MatCode"], m["Description"],
        m["Unit"], m["Qty"], m["InOut"], m["Supplier"],
        m["ไฟล์ต้นฉบับ"], m["นำเข้าเมื่อ"],
    ] for m in data["materials"]]
    append_to_sheet(ss, "Material Ledger", mat_rows)

    update_stock_summary(ss, data["materials"])

# ─── Main ─────────────────────────────────────────────────────────
def main():
    log.info("🚀 เริ่มระบบ ProChain Daily Report (Gmail Edition)")

    if not check_dead_key():
        return

    gmail = get_gmail_service()
    drive = get_drive_service()
    gc = get_gsheet_client()

    messages = fetch_today_emails(gmail)
    if not messages:
        log.info("📭 ไม่มีอีเมลใหม่วันนี้")
        return

    for msg_ref in messages:
        attachments, subject = fetch_attachments(gmail, msg_ref["id"])
        log.info(f"📧 ประมวลผล: {subject}")

        parts = subject.split("-")
        project_from_subject = parts[1].strip() if len(parts) > 1 else ""
        date_from_subject    = parts[2].strip() if len(parts) > 2 else ""

        for att in attachments:
            name = att["name"]
            content_b64 = att["content_b64"]
            log.info(f"📎 ประมวลผลไฟล์: {name}")

            data = parse_excel(content_b64, name)

            if not data["project_name"]:
                data["daily_log"]["ชื่อโครงการ"] = project_from_subject
                data["project_name"] = project_from_subject
            if not data["report_date"]:
                data["daily_log"]["วันที่"] = date_from_subject
                data["report_date"] = date_from_subject

            try:
                save_to_drive(drive, name, content_b64, data["project_name"], data["report_date"])
            except Exception as e:
                log.warning(f"⚠️ บันทึก Drive ไม่ได้: {e} — ดำเนินการต่อ")

            write_to_gsheet(gc, data)

    log.info("✅ เสร็จสิ้น")

if __name__ == "__main__":
    main()
